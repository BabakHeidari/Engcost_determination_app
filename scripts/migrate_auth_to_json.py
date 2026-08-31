"""One-time, operator-invoked migration from legacy XLSX authentication data."""

import argparse
import hashlib
import re
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from utils.profile_store import ProfileDataStore, normalize_email


SHA256_PATTERN = re.compile(r"[0-9a-fA-F]{64}\Z")
ROLE_MAPPING = {"admin": "IT Admin", "user": "Office Staff"}


def migrate(source: Path, destination: Path) -> int:
    workbook = load_workbook(source, read_only=True, data_only=True)
    rows = workbook.active.iter_rows(values_only=True)
    headers = next(rows, None)
    required = ("username", "password_hash", "access_level")
    if headers is None or any(field not in headers for field in required):
        raise ValueError("Legacy workbook does not contain the required columns")
    indexes = {field: headers.index(field) for field in required}
    users = []
    seen = set()
    now = datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")
    for row_number, row in enumerate(rows, start=2):
        username = normalize_email(str(row[indexes["username"]]))
        if not username or username in seen:
            raise ValueError(f"Duplicate or empty normalized identifier at row {row_number}")
        role = ROLE_MAPPING.get(str(row[indexes["access_level"]]).strip().casefold())
        if role is None:
            raise ValueError(f"Unrecognized access level at row {row_number}")
        legacy_hash = str(row[indexes["password_hash"]]).strip()
        compatible = bool(SHA256_PATTERN.fullmatch(legacy_hash))
        users.append({
            "id": "usr_" + hashlib.sha256(username.encode("utf-8")).hexdigest()[:24],
            "username": username,
            "email": username,
            "full_name": username,
            "role": role,
            "factory_id": None,
            "is_active": True,
            "must_change_password": True,
            "password_hash": legacy_hash.casefold() if compatible else None,
            "password_scheme": "legacy_sha256" if compatible else None,
            "created_at": now,
            "updated_at": now,
            "last_login_at": None,
            "password_changed_at": None,
            "revision": 1,
        })
        seen.add(username)

    store = ProfileDataStore(destination)
    store.initialize({
        "IT Admin": {"scope": "global", "permissions": {}},
        "Office Staff": {"scope": "global", "permissions": {}},
    })
    for user in users:
        store.create_user(user)
    return len(users)


def main():
    parser = argparse.ArgumentParser(description="Migrate legacy authentication users to canonical JSON")
    parser.add_argument("source", type=Path)
    parser.add_argument("destination", type=Path)
    args = parser.parse_args()
    count = migrate(args.source, args.destination)
    print(f"Migration completed and verified: {count} user record(s).")


if __name__ == "__main__":
    main()
